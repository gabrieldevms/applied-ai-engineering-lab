import csv
import hashlib
from collections.abc import Iterable
from io import BytesIO, StringIO
from pathlib import Path
from typing import Protocol
from docx import Document
from openpyxl import load_workbook
from ai_api.rag.exceptions import TextExtractionError
from ai_api.rag.schemas import (
    ExtractedTable,
    StructuredTableExtractionResponse,
)


class StructuredTableExtractor(Protocol):
    name: str
    supported_extensions: frozenset[str]

    def extract_tables(
        self,
        file_content: bytes,
        filename: str,
        content_type: str | None = None,
    ) -> StructuredTableExtractionResponse:
        """Extract structured tables from a supported file."""
        ...


def normalize_table_cell(value: object) -> str:
    if value is None:
        return ""

    return " ".join(
        str(value).replace("\r\n", "\n").replace("\r", "\n").strip().split()
    )


def get_file_extension(filename: str) -> str:
    cleaned_filename = filename.strip() or "uploaded-file"

    return Path(cleaned_filename).suffix.lower()


def build_table_id(
    filename: str,
    table_index: int,
    sheet_name: str | None = None,
) -> str:
    raw_value = f"{filename}:{table_index}:{sheet_name or ''}"
    digest = hashlib.sha256(
        raw_value.encode("utf-8")
    ).hexdigest()[:12]

    return f"table-{digest}"


def normalize_rows(rows: Iterable[Iterable[object]]) -> list[list[str]]:
    normalized_rows: list[list[str]] = []

    for row in rows:
        normalized_row = [
            normalize_table_cell(value)
            for value in row
        ]

        while normalized_row and not normalized_row[-1]:
            normalized_row.pop()

        if any(normalized_row):
            normalized_rows.append(normalized_row)

    return normalized_rows


def build_extracted_table(
    filename: str,
    rows: list[list[str]],
    table_index: int,
    sheet_name: str | None = None,
    metadata: dict | None = None,
) -> ExtractedTable:
    if not rows:
        raise TextExtractionError("Extracted tables are empty.")

    column_count = max(
        len(row)
        for row in rows
    )

    return ExtractedTable(
        table_id=build_table_id(
            filename=filename,
            table_index=table_index,
            sheet_name=sheet_name,
        ),
        source=filename,
        filename=filename,
        table_index=table_index,
        rows=rows,
        row_count=len(rows),
        column_count=column_count,
        sheet_name=sheet_name,
        metadata=metadata or {},
    )


class CSVStructuredTableExtractor:
    name = "csv-table"

    supported_extensions = frozenset(
        {
            ".csv",
        }
    )

    def extract_tables(
        self,
        file_content: bytes,
        filename: str,
        content_type: str | None = None,
    ) -> StructuredTableExtractionResponse:
        cleaned_filename = filename.strip() or "uploaded-file"
        extension = get_file_extension(cleaned_filename)

        if extension not in self.supported_extensions:
            raise TextExtractionError(
                f"Unsupported table extraction file type: {extension or 'unknown'}"
            )

        try:
            decoded_content = file_content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise TextExtractionError(
                "CSV content could not be decoded as UTF-8 text."
            ) from exc

        if not decoded_content.strip():
            raise TextExtractionError("Extracted tables are empty.")

        try:
            dialect = csv.Sniffer().sniff(decoded_content[:2048])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(
            StringIO(decoded_content),
            dialect=dialect,
        )

        rows = normalize_rows(reader)

        if not rows:
            raise TextExtractionError("Extracted tables are empty.")

        table = build_extracted_table(
            filename=cleaned_filename,
            rows=rows,
            table_index=0,
            metadata={
                "extension": extension,
                "extraction_method": self.name,
                "extractor": type(self).__name__,
                "delimiter": dialect.delimiter,
            },
        )

        return StructuredTableExtractionResponse(
            source=cleaned_filename,
            filename=cleaned_filename,
            content_type=content_type,
            tables=[
                table,
            ],
            table_count=1,
            metadata={
                "extension": extension,
                "extraction_method": self.name,
                "extractor": type(self).__name__,
                "total_rows": table.row_count,
                "max_columns": table.column_count,
            },
        )


class ExcelStructuredTableExtractor:
    name = "excel-table"

    supported_extensions = frozenset(
        {
            ".xlsx",
        }
    )

    def extract_tables(
        self,
        file_content: bytes,
        filename: str,
        content_type: str | None = None,
    ) -> StructuredTableExtractionResponse:
        cleaned_filename = filename.strip() or "uploaded-file"
        extension = get_file_extension(cleaned_filename)

        if extension not in self.supported_extensions:
            raise TextExtractionError(
                f"Unsupported table extraction file type: {extension or 'unknown'}"
            )

        try:
            workbook = load_workbook(
                BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise TextExtractionError(
                "Excel tables could not be extracted."
            ) from exc

        try:
            tables: list[ExtractedTable] = []

            for worksheet in workbook.worksheets:
                rows = normalize_rows(
                    worksheet.iter_rows(values_only=True)
                )

                if not rows:
                    continue

                tables.append(
                    build_extracted_table(
                        filename=cleaned_filename,
                        rows=rows,
                        table_index=len(tables),
                        sheet_name=worksheet.title,
                        metadata={
                            "extension": extension,
                            "extraction_method": self.name,
                            "extractor": type(self).__name__,
                            "sheet_name": worksheet.title,
                        },
                    )
                )

            if not tables:
                raise TextExtractionError("Extracted tables are empty.")

            return StructuredTableExtractionResponse(
                source=cleaned_filename,
                filename=cleaned_filename,
                content_type=content_type,
                tables=tables,
                table_count=len(tables),
                metadata={
                    "extension": extension,
                    "extraction_method": self.name,
                    "extractor": type(self).__name__,
                    "sheet_count": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                    "total_rows": sum(
                        table.row_count
                        for table in tables
                    ),
                    "max_columns": max(
                        table.column_count
                        for table in tables
                    ),
                },
            )
        finally:
            workbook.close()


class DOCXStructuredTableExtractor:
    name = "docx-table"

    supported_extensions = frozenset(
        {
            ".docx",
        }
    )

    def extract_tables(
        self,
        file_content: bytes,
        filename: str,
        content_type: str | None = None,
    ) -> StructuredTableExtractionResponse:
        cleaned_filename = filename.strip() or "uploaded-file"
        extension = get_file_extension(cleaned_filename)

        if extension not in self.supported_extensions:
            raise TextExtractionError(
                f"Unsupported table extraction file type: {extension or 'unknown'}"
            )

        try:
            document = Document(BytesIO(file_content))
        except Exception as exc:
            raise TextExtractionError(
                "DOCX tables could not be extracted."
            ) from exc

        tables: list[ExtractedTable] = []

        for document_table in document.tables:
            rows = normalize_rows(
                [
                    [
                        cell.text
                        for cell in row.cells
                    ]
                    for row in document_table.rows
                ]
            )

            if not rows:
                continue

            tables.append(
                build_extracted_table(
                    filename=cleaned_filename,
                    rows=rows,
                    table_index=len(tables),
                    metadata={
                        "extension": extension,
                        "extraction_method": self.name,
                        "extractor": type(self).__name__,
                    },
                )
            )

        if not tables:
            raise TextExtractionError("Extracted tables are empty.")

        return StructuredTableExtractionResponse(
            source=cleaned_filename,
            filename=cleaned_filename,
            content_type=content_type,
            tables=tables,
            table_count=len(tables),
            metadata={
                "extension": extension,
                "extraction_method": self.name,
                "extractor": type(self).__name__,
                "table_count": len(tables),
                "total_rows": sum(
                    table.row_count
                    for table in tables
                ),
                "max_columns": max(
                    table.column_count
                    for table in tables
                ),
            },
        )


class StructuredTableExtractorRegistry:
    def __init__(
        self,
        extractors: Iterable[StructuredTableExtractor] | None = None,
    ) -> None:
        selected_extractors = (
            [
                CSVStructuredTableExtractor(),
                ExcelStructuredTableExtractor(),
                DOCXStructuredTableExtractor(),
            ]
            if extractors is None
            else list(extractors)
        )

        self._extractors_by_extension: dict[str, StructuredTableExtractor] = {}

        for extractor in selected_extractors:
            self.register(extractor)

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        return tuple(sorted(self._extractors_by_extension))

    def register(self, extractor: StructuredTableExtractor) -> None:
        if not extractor.supported_extensions:
            raise ValueError(
                "Structured table extractor must define at least one supported extension."
            )

        normalized_extensions = {
            self._normalize_extension(extension)
            for extension in extractor.supported_extensions
        }

        for extension in normalized_extensions:
            if extension in self._extractors_by_extension:
                raise ValueError(
                    "Structured table extractor already registered for extension: "
                    f"{extension}"
                )

        for extension in normalized_extensions:
            self._extractors_by_extension[extension] = extractor

    def get_for_filename(self, filename: str) -> StructuredTableExtractor:
        cleaned_filename = filename.strip() or "uploaded-file"
        extension = get_file_extension(cleaned_filename)

        return self.get_for_extension(extension)

    def get_for_extension(
        self,
        extension: str,
    ) -> StructuredTableExtractor:
        normalized_extension = self._normalize_extension(
            extension,
            allow_empty=True,
        )

        extractor = self._extractors_by_extension.get(
            normalized_extension
        )

        if extractor is None:
            raise TextExtractionError(
                "Unsupported table extraction file type: "
                f"{normalized_extension or 'unknown'}"
            )

        return extractor

    def _normalize_extension(
        self,
        extension: str,
        allow_empty: bool = False,
    ) -> str:
        normalized_extension = extension.strip().lower()

        if not normalized_extension:
            if allow_empty:
                return ""

            raise ValueError("File extension cannot be blank.")

        if not normalized_extension.startswith("."):
            normalized_extension = f".{normalized_extension}"

        return normalized_extension


class TableExtractionService:
    def __init__(
        self,
        extractor_registry: StructuredTableExtractorRegistry | None = None,
    ) -> None:
        self.extractor_registry = (
            extractor_registry
            if extractor_registry is not None
            else StructuredTableExtractorRegistry()
        )

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        return self.extractor_registry.supported_extensions

    def extract_from_bytes(
        self,
        file_content: bytes,
        filename: str,
        content_type: str | None = None,
    ) -> StructuredTableExtractionResponse:
        cleaned_filename = filename.strip() or "uploaded-file"

        extractor = self.extractor_registry.get_for_filename(
            cleaned_filename
        )

        return extractor.extract_tables(
            file_content=file_content,
            filename=cleaned_filename,
            content_type=content_type,
        )
